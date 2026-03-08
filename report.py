import os
import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
import datetime

# ──────────────────────────────────────────────────────────────────────────────
# Cluster ordering
# ──────────────────────────────────────────────────────────────────────────────
CLUSTER_ORDER = [
    'Central', 'Eastern', 'North Eastern',
    'Northern', 'South Eastern', 'Southern',
]

_CLUSTER_ALIASES = {
    'central cluster': 'Central', 'central': 'Central',
    'eastern cluster': 'Eastern', 'eastern': 'Eastern',
    'ne cluster': 'North Eastern', 'northeast cluster': 'North Eastern',
    'north eastern cluster': 'North Eastern', 'north eastern': 'North Eastern',
    'northeastern': 'North Eastern',
    'northern cluster': 'Northern', 'northern': 'Northern',
    'se cluster': 'South Eastern', 'southeast cluster': 'South Eastern',
    'south eastern cluster': 'South Eastern', 'south eastern': 'South Eastern',
    'southeastern': 'South Eastern',
    'southern cluster': 'Southern', 'southern': 'Southern',
}

def _match_cluster(name):
    if not isinstance(name, str) or not name.strip():
        return None
    key = name.strip().lower()
    if key in _CLUSTER_ALIASES:
        return _CLUSTER_ALIASES[key]
    for alias, canonical in _CLUSTER_ALIASES.items():
        if alias in key:
            return canonical
    return None

# ──────────────────────────────────────────────────────────────────────────────
# Header Validation
# ──────────────────────────────────────────────────────────────────────────────
REQUIRED_COLUMNS = {
    'Cluster Name':     ['Cluster Name', 'Cluster_Name', 'cluster name'],
    'Work Status':      ['Work Status', 'work status', 'Work_Status'],
    'Action Type':      ['Action Type', 'Action_Type', 'action type'],
    'TT Create Time':   ['TT CREATE TIME', 'TT_CREATE_TIME', 'TT Create Time'],
    'TT Recovery Time': ['TT Recovery Time', 'TT_RECOVERY_TIME'],
    'Fault Level':      ['FAULT_LEVEL', 'TT Type', 'TT_TYPE'],
    'SITE_ID':          ['SITE_ID', 'Site ID', 'SiteID'],
}

def _validate_headers(df_cols):
    missing = []
    for concept, aliases in REQUIRED_COLUMNS.items():
        if not any(a in df_cols for a in aliases):
            missing.append(concept)
    if missing:
        raise ValueError(
            f"Missing required columns: {', '.join(missing)}.\n"
            f"Please ensure your Excel file contains these columns and try again."
        )

# ──────────────────────────────────────────────────────────────────────────────
# Time helpers
# ──────────────────────────────────────────────────────────────────────────────
def _td_to_excel(td):
    """Convert timedelta → Excel fraction-of-day float."""
    if pd.isnull(td):
        return None
    return td.total_seconds() / 86400.0

def _safe_dt(v):
    try:
        return pd.to_datetime(v)
    except Exception:
        return pd.NaT

def _safe_date(v):
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None

# ──────────────────────────────────────────────────────────────────────────────
# Styling constants
# ──────────────────────────────────────────────────────────────────────────────
# Fills
HEADER_FILL  = PatternFill('solid', fgColor='1F3864')   # deep navy
SUBHEAD_FILL = PatternFill('solid', fgColor='2E75B6')   # medium blue
TOTAL_FILL   = PatternFill('solid', fgColor='BDD7EE')   # pale blue
ALT_FILL     = PatternFill('solid', fgColor='DEEAF1')   # very light blue
WHITE_FILL   = PatternFill('solid', fgColor='FFFFFF')
TITLE_FILL   = PatternFill('solid', fgColor='0D2137')   # darkest navy
GREEN_FILL   = PatternFill('solid', fgColor='C6EFCE')   # light green
YELLOW_FILL  = PatternFill('solid', fgColor='FFEB9C')   # light yellow
RED_FILL     = PatternFill('solid', fgColor='FFC7CE')   # light red

# Fonts
HEADER_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
SUBHEAD_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
TOTAL_FONT   = Font(name='Calibri', bold=True, color='1F3864', size=10)
NORMAL_FONT  = Font(name='Calibri', size=10)
TITLE_FONT   = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
SMALL_FONT   = Font(name='Calibri', size=9)

# Borders
_THIN = Side(style='thin', color='B8CCE4')
_MED  = Side(style='medium', color='1F3864')
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
MED_BORDER  = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)

# Alignments
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

# Number formats
TIME_FMT  = '[h]:mm'
PCT_FMT   = '0.00%'
INT_FMT   = '#,##0'


def _sc(ws, row, col, value, fill=None, font=None, align=None, border=None, fmt=None):
    """Style cell helper."""
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if align:  c.alignment = align
    if border: c.border = border
    if fmt:    c.number_format = fmt
    return c


def _merge_title(ws, row, col1, col2, text, fill=HEADER_FILL, font=None):
    font = font or HEADER_FONT
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    _sc(ws, row, col1, text, fill, font, CENTER)


# ──────────────────────────────────────────────────────────────────────────────
# Main report generator
# ──────────────────────────────────────────────────────────────────────────────
def generate_report(filepath: str) -> bytes:
    """Read raw data Excel and return bytes of the generated report workbook."""

    # 1. Load & validate
    df_raw = pd.read_excel(filepath, sheet_name=0)
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    _validate_headers(df_raw.columns)

    # Keep numeric/datetime columns native, string-ify only for label matching
    df_str = df_raw.astype(str).replace('nan', np.nan)

    # 2. Resolve column names
    def _find(aliases):
        return next((c for c in df_str.columns if c in aliases), None)

    cluster_col = _find(REQUIRED_COLUMNS['Cluster Name'])
    ws_col      = _find(REQUIRED_COLUMNS['Work Status'])
    tt_col      = _find(REQUIRED_COLUMNS['Fault Level'])
    action_col  = _find(REQUIRED_COLUMNS['Action Type'])
    subtype_col = _find(['Sub_Type', 'Sub Type', 'SubType'])
    fo_col      = next((c for c in df_str.columns if 'ticket update' in c.lower() or 'fo movement' in c.lower()), None)
    pts_col     = next((c for c in df_str.columns if 'parent ticket status' in c.lower()), None)
    site_col    = _find(REQUIRED_COLUMNS['SITE_ID'])
    date_col    = next((c for c in df_str.columns if c.strip().lower() == 'date'), None)
    ttc_col     = _find(REQUIRED_COLUMNS['TT Create Time'])
    ttr_col     = _find(REQUIRED_COLUMNS['TT Recovery Time'])

    # 3. Build working columns on df_str
    df_str['_Cluster']     = df_str[cluster_col].apply(_match_cluster)
    df_str['_WorkStatus']  = df_str[ws_col].str.strip()

    def _map_tt(v):
        u = str(v).strip().upper()
        if 'P1' in u or 'CRITICAL' in u: return 'P1'
        if 'P2' in u or 'MAJOR'    in u: return 'P2'
        if 'P3' in u or 'MINOR'    in u: return 'P3'
        return u
    df_str['_TTType']      = df_str[tt_col].apply(_map_tt)
    df_str['_ActionType']  = df_str[action_col].str.strip() if action_col else 'Unknown'
    df_str['_SubType']     = df_str[subtype_col].str.strip() if subtype_col else df_str['_ActionType']

    def _map_fo(v):
        v = str(v).strip().lower()
        if 'multi ticket' in v or 'multiple tt' in v: return 'Multiple TT, Single Movement'
        if ('single ticket' in v and ('multi site' in v or 'multi movement' in v)) \
                or 'single tt, multiple movement' in v: return 'Single TT, Multiple Movement'
        return 'Single TT, Single Movement'
    df_str['_FO']          = df_str[fo_col].apply(_map_fo) if fo_col else 'Single TT, Single Movement'
    df_str['_SiteID']      = df_str[site_col].str.strip() if site_col else ''
    df_str['_ParentStatus']= df_str[pts_col].str.strip() if pts_col else ''

    # MTTR from original (datetime-aware) frame
    df_raw['_TTCreate']   = df_raw[ttc_col].apply(_safe_dt)
    df_raw['_TTRecovery'] = df_raw[ttr_col].apply(_safe_dt)
    df_raw['_E2EMTTR']   = df_raw['_TTRecovery'] - df_raw['_TTCreate']

    # Copy MTTR into str frame
    df_str['_E2EMTTR'] = df_raw['_E2EMTTR']

    # 4. Unique TT flag
    uniq_col = next((c for c in df_str.columns if c.strip() == 'Unique TT'), None)
    if uniq_col:
        df_str['_UniqTT'] = pd.to_numeric(df_str[uniq_col], errors='coerce').fillna(0).astype(int)
    elif 'PARENT_TICKET_ID' in df_str.columns:
        df_str['_UniqTT'] = (~df_str['PARENT_TICKET_ID'].duplicated(keep='first')).astype(int)
    else:
        df_str['_UniqTT'] = 1

    # 5. Site recurrence helper
    if site_col:
        raw_sites = df_str[['_SiteID']].copy()
        if date_col:
            raw_sites['_Date'] = df_raw[date_col].apply(_safe_date)
            s_map = raw_sites.dropna().groupby('_SiteID')['_Date'].nunique().to_dict()
        else:
            s_map = raw_sites['_SiteID'].value_counts().to_dict()
        df_str['_SiteRepeat'] = df_str['_SiteID'].map(s_map).fillna(0).astype(int)
    else:
        df_str['_SiteRepeat'] = 0

    # 6. Working subsets
    df      = df_str[df_str['_Cluster'].notna()].copy()
    df_uniq = df[df['_UniqTT'] == 1].copy()

    # 6. Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Note: raw_refs (Named Ranges) are defined in _build_raw_sheet
    # We call it first to get the references, but we will move it to the end later
    raw_refs = _build_raw_sheet(wb, df_str)
    _build_summary_sheet(wb, df, df_uniq, date_col, raw_refs)
    _build_mttr_sheet(wb, df_uniq, raw_refs)

    # Move Raw Data to the end
    raw_sheet = wb['Raw Data']
    # Removing and re-adding is a simple way to move it to the end in openpyxl
    # But better to just reorder the sheet list if possible, or build it logically.
    # Actually, openpyxl appends by default. If I want it last, I should add Summary and MTTR FIRST.
    # But they need the Named Ranges from Raw Data.
    # I will move the sheet index.
    idx = len(wb.sheetnames) - 1
    wb._sheets = [wb._sheets[1], wb._sheets[2], wb._sheets[0]]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────────────────────────────────────
# Helper: Write a single MTTR pivot table
# Returns (first_data_row, last_data_row) for chart use
# ──────────────────────────────────────────────────────────────────────────────
def _write_mttr_table(ws, df_prio, start_row, prio_label, prio_code, raw_refs):
    """Write one priority's MTTR table using Named Ranges."""
    col_headers = CLUSTER_ORDER + ['Grand Total']
    n_clusters  = len(CLUSTER_ORDER)

    _merge_title(ws, start_row, 1, n_clusters + 2, f'Average E2E MTTR — {prio_label}', HEADER_FILL)
    _sc(ws, start_row + 1, 1, 'RC / On-Site Action', SUBHEAD_FILL, SUBHEAD_FONT, LEFT, THIN_BORDER)
    for ci, ch in enumerate(col_headers, start=2):
        _sc(ws, start_row + 1, ci, ch, SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)

    data_start = start_row + 2
    subtypes = sorted(df_prio['_SubType'].dropna().unique()) if len(df_prio) else []

    if not subtypes:
        _sc(ws, data_start, 1, '— No data for this priority —', WHITE_FILL, NORMAL_FONT, LEFT, THIN_BORDER)
        return data_start, data_start

    row_idx = data_start
    for st in subtypes:
        fill = ALT_FILL if (row_idx % 2 == 0) else WHITE_FILL
        _sc(ws, row_idx, 1, st, fill, NORMAL_FONT, LEFT, THIN_BORDER)
        
        # Column formulas for each cluster
        for ci, cl in enumerate(CLUSTER_ORDER, start=2):
            f = f'=IFERROR(AVERAGEIFS(Raw_MTTR, Raw_Cluster, "{cl}", Raw_SubType, "{st}", Raw_Type, "{prio_code}"), "")'
            c = ws.cell(row=row_idx, column=ci, value=f)
            c.number_format = TIME_FMT
            c.fill = fill; c.font = NORMAL_FONT; c.alignment = CENTER; c.border = THIN_BORDER
        
        # Row Grand total
        f_gt = f'=IFERROR(AVERAGEIFS(Raw_MTTR, Raw_SubType, "{st}", Raw_Type, "{prio_code}"), "")'
        c = ws.cell(row=row_idx, column=n_clusters + 2, value=f_gt)
        c.number_format = TIME_FMT; c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        c.alignment = CENTER; c.border = THIN_BORDER
        row_idx += 1

    # Grand total row
    gt_row = row_idx
    _sc(ws, gt_row, 1, 'Grand Total', TOTAL_FILL, TOTAL_FONT, LEFT, THIN_BORDER)
    for ci, cl in enumerate(CLUSTER_ORDER, start=2):
        f = f'=IFERROR(AVERAGEIFS(Raw_MTTR, Raw_Cluster, "{cl}", Raw_Type, "{prio_code}"), "")'
        c = ws.cell(row=gt_row, column=ci, value=f)
        c.number_format = TIME_FMT; c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        c.alignment = CENTER; c.border = THIN_BORDER

    f_all = f'=IFERROR(AVERAGEIFS(Raw_MTTR, Raw_Type, "{prio_code}"), "")'
    c = ws.cell(row=gt_row, column=n_clusters + 2, value=f_all)
    c.number_format = TIME_FMT; c.fill = TOTAL_FILL; c.font = TOTAL_FONT
    c.alignment = CENTER; c.border = THIN_BORDER

    return data_start, gt_row


# ──────────────────────────────────────────────────────────────────────────────
# MTTR Sheet
# ──────────────────────────────────────────────────────────────────────────────
def _build_mttr_sheet(wb: Workbook, df_uniq: pd.DataFrame, raw_refs):
    ws = wb.create_sheet('MTTR')
    ws.sheet_view.showGridLines = False

    # Big title
    _merge_title(ws, 1, 1, 10, 'Average E2E MTTR Analysis', TITLE_FILL, TITLE_FONT)

    current_row = 3
    for prio in ['P1', 'P2', 'P3']:
        df_prio = df_uniq[(df_uniq['_TTType'] == prio) & df_uniq['_E2EMTTR'].notna()].copy()
        data_start, gt_row = _write_mttr_table(ws, df_prio, current_row, f'{prio} Tickets', prio, raw_refs)
        current_row = gt_row + 3   # gap between priority sections

    # Column widths
    ws.column_dimensions['A'].width = 30
    for ci in range(2, len(CLUSTER_ORDER) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = 'B3'


# ──────────────────────────────────────────────────────────────────────────────
# Summary Sheet
# ──────────────────────────────────────────────────────────────────────────────
def _build_summary_sheet(wb: Workbook, df: pd.DataFrame, df_uniq: pd.DataFrame, date_col, raw_refs):
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    n_cl = len(CLUSTER_ORDER)

    # ── Section 1: Big title ──────────────────────────────────────────────────
    _merge_title(ws, 1, 1, 18, 'Daily Activity Report — Summary', TITLE_FILL, TITLE_FONT)
    ws.row_dimensions[1].height = 24

    # ── Section 2: Work Status / FO / MTTR / Recurrence ──────────────────────
    top_start = 3

    # Grouped headers row
    _merge_title(ws, top_start, 2, 5,  'Work Status',           HEADER_FILL)
    _merge_title(ws, top_start, 6, 9,  'FO Movement | Unique TT', HEADER_FILL)
    _sc(ws, top_start, 10, 'FTR',    SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)
    _sc(ws, top_start, 11, 'E2E MTTR', SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)
    _merge_title(ws, top_start, 12, 17, 'Site Recurrence',       HEADER_FILL)
    _sc(ws, top_start, 1,  'Cluster Name', HEADER_FILL, HEADER_FONT, LEFT, THIN_BORDER)
    ws.row_dimensions[top_start].height = 18

    # Sub-headers
    sub = ['Cluster Name',
           'Closed', 'Dependency', 'WIP', 'Total (WS)',
           'Multi TT, Single Move', 'Single TT, Multi Move', 'Single TT, Single Move', 'Total (FO)',
           'FTR', 'E2E MTTR',
           '1x', '2x', '3x', '4x', '5x+', 'Total (Sites)', 'Recurrence %']
    for ci, h in enumerate(sub, start=1):
        _sc(ws, top_start + 1, ci, h, SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)
    ws.row_dimensions[top_start + 1].height = 30

    data_start_row = top_start + 2
    clusters_all   = CLUSTER_ORDER + ['Grand Total']

    for ri, cl in enumerate(clusters_all):
        r       = data_start_row + ri
        is_gt   = (cl == 'Grand Total')
        fill    = TOTAL_FILL if is_gt else (ALT_FILL if ri % 2 == 0 else WHITE_FILL)
        font    = TOTAL_FONT if is_gt else NORMAL_FONT

        # A: Cluster name
        _sc(ws, r, 1, cl, fill, font, LEFT, THIN_BORDER)

        # B-E: Work Status
        def _cf(status):
            if is_gt: return f'=COUNTIFS(Raw_Status, "{status}", Raw_Uniq, 1)'
            return f'=COUNTIFS(Raw_Cluster, $A{r}, Raw_Status, "{status}", Raw_Uniq, 1)'

        ws.cell(row=r, column=2, value=_cf('Closed')).fill=fill; ws.cell(row=r, column=2).font=font; ws.cell(row=r, column=2).border=THIN_BORDER
        ws.cell(row=r, column=3, value=_cf('Dependency')).fill=fill; ws.cell(row=r, column=3).font=font; ws.cell(row=r, column=3).border=THIN_BORDER
        ws.cell(row=r, column=4, value=_cf('WIP')).fill=fill; ws.cell(row=r, column=4).font=font; ws.cell(row=r, column=4).border=THIN_BORDER
        
        if is_gt:
            ws.cell(row=r, column=5, value=f'=COUNTIFS(Raw_Uniq, 1)').fill=fill
        else:
            ws.cell(row=r, column=5, value=f'=COUNTIFS(Raw_Cluster, $A{r}, Raw_Uniq, 1)').fill=fill
        ws.cell(row=r, column=5).font=font; ws.cell(row=r, column=5).border=THIN_BORDER

        # F-I: FO Movement
        def _cfo(label):
            if is_gt: return f'=COUNTIFS(Raw_FO, "{label}", Raw_Uniq, 1)'
            return f'=COUNTIFS(Raw_Cluster, $A{r}, Raw_FO, "{label}", Raw_Uniq, 1)'

        ws.cell(row=r, column=6, value=_cfo('Multiple TT, Single Movement')).fill=fill; ws.cell(row=r, column=6).font=font; ws.cell(row=r, column=6).border=THIN_BORDER
        ws.cell(row=r, column=7, value=_cfo('Single TT, Multiple Movement')).fill=fill; ws.cell(row=r, column=7).font=font; ws.cell(row=r, column=7).border=THIN_BORDER
        ws.cell(row=r, column=8, value=_cfo('Single TT, Single Movement')).fill=fill; ws.cell(row=r, column=8).font=font; ws.cell(row=r, column=8).border=THIN_BORDER
        ws.cell(row=r, column=9, value=f'=E{r}').fill=fill; ws.cell(row=r, column=9).font=font; ws.cell(row=r, column=9).border=THIN_BORDER

        # J: FTR %
        ws.cell(row=r, column=10, value=f'=IFERROR((F{r}+H{r})/I{r}, "")').number_format=PCT_FMT
        ws.cell(row=r, column=10).fill=fill; ws.cell(row=r, column=10).font=font; ws.cell(row=r, column=10).border=THIN_BORDER

        # K: E2E MTTR
        if is_gt:
            f_mttr = f'=IFERROR(AVERAGEIFS(Raw_MTTR, Raw_Uniq, 1), "")'
        else:
            f_mttr = f'=IFERROR(AVERAGEIFS(Raw_MTTR, Raw_Cluster, $A{r}, Raw_Uniq, 1), "")'
        c_mttr = ws.cell(row=r, column=11, value=f_mttr)
        c_mttr.number_format = TIME_FMT; c_mttr.fill=fill; c_mttr.font=font; c_mttr.border=THIN_BORDER

        # L-R: Site Recurrence
        def _csr(val, op='='):
            if is_gt: return f'=COUNTIFS(Raw_Repeat, "{op}{val}")'
            return f'=COUNTIFS(Raw_Cluster, $A{r}, Raw_Repeat, "{op}{val}")'

        ws.cell(row=r, column=12, value=_csr(1)).fill=fill; ws.cell(row=r, column=12).font=font; ws.cell(row=r, column=12).border=THIN_BORDER
        ws.cell(row=r, column=13, value=_csr(2)).fill=fill; ws.cell(row=r, column=13).font=font; ws.cell(row=r, column=13).border=THIN_BORDER
        ws.cell(row=r, column=14, value=_csr(3)).fill=fill; ws.cell(row=r, column=14).font=font; ws.cell(row=r, column=14).border=THIN_BORDER
        ws.cell(row=r, column=15, value=_csr(4)).fill=fill; ws.cell(row=r, column=15).font=font; ws.cell(row=r, column=15).border=THIN_BORDER
        ws.cell(row=r, column=16, value=_csr(5, '>=')).fill=fill; ws.cell(row=r, column=16).font=font; ws.cell(row=r, column=16).border=THIN_BORDER
        
        ws.cell(row=r, column=17, value=f'=SUM(L{r}:P{r})').fill=fill; ws.cell(row=r, column=17).font=font; ws.cell(row=r, column=17).border=THIN_BORDER
        ws.cell(row=r, column=18, value=f'=IFERROR((M{r}+N{r}+O{r}+P{r})/Q{r}, "")').number_format=PCT_FMT
        ws.cell(row=r, column=18).fill=fill; ws.cell(row=r, column=18).font=font; ws.cell(row=r, column=18).border=THIN_BORDER

        ws.row_dimensions[r].height = 16

    top_end_row = data_start_row + len(clusters_all)

    # ── Section 3: RC / On-Site Findings ─────────────────────────────────────
    rc_hdr_row  = top_end_row + 2
    rc_data_col_end = n_cl + 2

    _merge_title(ws, rc_hdr_row, 1, rc_data_col_end + 1, 'RC / On-Site Findings', HEADER_FILL)
    rc_sub_row  = rc_hdr_row + 1
    _sc(ws, rc_sub_row, 1, 'Action Type', SUBHEAD_FILL, SUBHEAD_FONT, LEFT,   THIN_BORDER)
    for ci, cl in enumerate(CLUSTER_ORDER, start=2):
        _sc(ws, rc_sub_row, ci, cl, SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)
    _sc(ws, rc_sub_row, n_cl + 2, 'Grand Total',  SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)
    _sc(ws, rc_sub_row, n_cl + 3, '% of Total',   SUBHEAD_FILL, SUBHEAD_FONT, CENTER, THIN_BORDER)

    action_types  = sorted(df_uniq['_ActionType'].dropna().unique())
    rc_row        = rc_sub_row + 1
    rc_data_start = rc_row

    for at in action_types:
        fill = ALT_FILL if (rc_row % 2 == 0) else WHITE_FILL
        _sc(ws, rc_row, 1, at, fill, NORMAL_FONT, LEFT, THIN_BORDER)
        
        for ci, cl in enumerate(CLUSTER_ORDER, start=2):
            f = f'=COUNTIFS(Raw_Cluster, "{cl}", Raw_Action, "{at}", Raw_Uniq, 1)'
            _sc(ws, rc_row, ci, f, fill, NORMAL_FONT, CENTER, THIN_BORDER, INT_FMT)

        let_end = get_column_letter(n_cl + 1)
        _sc(ws, rc_row, n_cl + 2, f'=SUM(B{rc_row}:{let_end}{rc_row})', TOTAL_FILL, TOTAL_FONT, CENTER, THIN_BORDER, INT_FMT)
        
        # % Calculation
        gt_col_let = get_column_letter(n_cl + 2)
        _sc(ws, rc_row, n_cl + 3, f'={gt_col_let}{rc_row}/{gt_col_let}$GT_ROW$', fill, NORMAL_FONT, CENTER, THIN_BORDER, PCT_FMT)
        rc_row += 1

    # RC Grand Total
    rc_gt_row = rc_row
    _sc(ws, rc_gt_row, 1, 'Grand Total', TOTAL_FILL, TOTAL_FONT, LEFT, THIN_BORDER)
    for ci, cl in enumerate(CLUSTER_ORDER, start=2):
        let = get_column_letter(ci)
        _sc(ws, rc_gt_row, ci, f'=SUM({let}{rc_data_start}:{let}{rc_gt_row-1})', TOTAL_FILL, TOTAL_FONT, CENTER, THIN_BORDER, INT_FMT)
    
    _sc(ws, rc_gt_row, n_cl+2, f'=SUM({gt_col_let}{rc_data_start}:{gt_col_let}{rc_gt_row-1})', TOTAL_FILL, TOTAL_FONT, CENTER, THIN_BORDER, INT_FMT)
    _sc(ws, rc_gt_row, n_cl+3, 1.0, TOTAL_FILL, TOTAL_FONT, CENTER, THIN_BORDER, PCT_FMT)

    # Fix placeholders for %
    for r_fix in range(rc_data_start, rc_gt_row):
        cell = ws.cell(row=r_fix, column=n_cl+3)
        if isinstance(cell.value, str):
            cell.value = cell.value.replace('$GT_ROW$', str(rc_gt_row))

    # ── Section 4: MTTR Analysis by Priority (P1, P2, P3) ────────────────────
    prio_start_row = rc_gt_row + 2
    for prio in ['P1', 'P2', 'P3']:
        df_prio = df_uniq[(df_uniq['_TTType'] == prio) & df_uniq['_E2EMTTR'].notna()].copy()
        _, gt_row = _write_mttr_table(ws, df_prio, prio_start_row, f'{prio} Tickets', prio, raw_refs)
        prio_start_row = gt_row + 3

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    for ci in range(2, 20):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = 'B3'


# ──────────────────────────────────────────────────────────────────────────────
# Charts Sheet — all charts in a dedicated sheet, no overlap
# ──────────────────────────────────────────────────────────────────────────────
def _build_charts_sheet(wb: Workbook, df: pd.DataFrame, df_uniq: pd.DataFrame):
    """Write a helper data table and charts into a dedicated Charts sheet."""
    ws = wb.create_sheet('Charts')
    ws.sheet_view.showGridLines = False

    _merge_title(ws, 1, 1, 10, 'Visual Analytics Dashboard', TITLE_FILL, TITLE_FONT)
    ws.row_dimensions[1].height = 22

    # ── (A) Write helper data for charts in cols N onwards (hidden area) ──────
    DATA_COL_START = 20   # Col T — off-screen data helper

    # A1 – Work Status per cluster
    ws.cell(row=2, column=DATA_COL_START, value='Cluster').font = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 1, value='Closed').font = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 2, value='Dependency').font = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 3, value='WIP').font = SUBHEAD_FONT

    for ri, cl in enumerate(CLUSTER_ORDER, start=3):
        sub   = df_uniq[df_uniq['_Cluster'] == cl]
        cnt   = sub['_WorkStatus'].value_counts()
        ws.cell(row=ri, column=DATA_COL_START,     value=cl)
        ws.cell(row=ri, column=DATA_COL_START + 1, value=int(cnt.get('Closed',     0)))
        ws.cell(row=ri, column=DATA_COL_START + 2, value=int(cnt.get('Dependency', 0)))
        ws.cell(row=ri, column=DATA_COL_START + 3, value=int(cnt.get('WIP',        0)))
    ws_data_end = 3 + len(CLUSTER_ORDER) - 1

    # A2 – RC breakdown (action type vs grand total)
    action_types = sorted(df_uniq['_ActionType'].dropna().unique())
    ws.cell(row=2, column=DATA_COL_START + 6, value='Action Type').font = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 7, value='Count').font = SUBHEAD_FONT
    for ri, at in enumerate(action_types, start=3):
        cnt = int((df_uniq['_ActionType'] == at).sum())
        ws.cell(row=ri, column=DATA_COL_START + 6, value=at)
        ws.cell(row=ri, column=DATA_COL_START + 7, value=cnt)
    rc_data_end = 3 + len(action_types) - 1

    # A3 – Priority breakdown
    ws.cell(row=2, column=DATA_COL_START + 10, value='Priority').font = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 11, value='Count').font = SUBHEAD_FONT
    for ri, p in enumerate(['P1', 'P2', 'P3'], start=3):
        cnt = int((df_uniq['_TTType'] == p).sum())
        ws.cell(row=ri, column=DATA_COL_START + 10, value=p)
        ws.cell(row=ri, column=DATA_COL_START + 11, value=cnt)

    # A4 – Site recurrence summary per cluster
    ws.cell(row=2, column=DATA_COL_START + 14, value='Cluster').font = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 15, value='1x').font       = SUBHEAD_FONT
    ws.cell(row=2, column=DATA_COL_START + 16, value='2x+').font      = SUBHEAD_FONT
    for ri, cl in enumerate(CLUSTER_ORDER, start=3):
        sub  = df[df['_Cluster'] == cl]
        site_cnt = sub.groupby('_SiteID').size()
        s1   = int((site_cnt == 1).sum())
        s2p  = int((site_cnt >= 2).sum())
        ws.cell(row=ri, column=DATA_COL_START + 14, value=cl)
        ws.cell(row=ri, column=DATA_COL_START + 15, value=s1)
        ws.cell(row=ri, column=DATA_COL_START + 16, value=s2p)

    # ── (B) Build Charts ──────────────────────────────────────────────────────
    chart_row = 3   # anchor row for first chart column

    # --- Chart 1: Work Status Stacked Bar ---
    c1 = BarChart()
    c1.type    = 'col'
    c1.grouping = 'stacked'
    c1.overlap  = 100
    c1.title   = 'Work Status by Cluster'
    c1.y_axis.title = 'Ticket Count'
    c1.x_axis.title = 'Cluster'
    c1.style   = 10
    c1_data    = Reference(ws, min_col=DATA_COL_START + 1, max_col=DATA_COL_START + 3,
                           min_row=2, max_row=ws_data_end)
    c1_cats    = Reference(ws, min_col=DATA_COL_START, min_row=3, max_row=ws_data_end)
    c1.add_data(c1_data, titles_from_data=True)
    c1.set_categories(c1_cats)
    c1.width = 20; c1.height = 14
    ws.add_chart(c1, 'B3')

    # --- Chart 2: RC Breakdown Pie ---
    c2 = PieChart()
    c2.title = 'RC / On-Site Findings'
    c2.style = 10
    c2_data  = Reference(ws, min_col=DATA_COL_START + 7, min_row=2, max_row=rc_data_end)
    c2_cats  = Reference(ws, min_col=DATA_COL_START + 6, min_row=3, max_row=rc_data_end)
    c2.add_data(c2_data, titles_from_data=True)
    c2.set_categories(c2_cats)
    c2.dataLabels               = DataLabelList()
    c2.dataLabels.showPercent   = True
    c2.dataLabels.showCatName   = False
    c2.dataLabels.showVal       = False
    c2.width = 17; c2.height = 14
    ws.add_chart(c2, 'L3')

    # --- Chart 3: Priority Distribution Pie ---
    c3 = PieChart()
    c3.title = 'Ticket Priority Breakdown'
    c3.style = 10
    c3_data  = Reference(ws, min_col=DATA_COL_START + 11, min_row=2, max_row=5)
    c3_cats  = Reference(ws, min_col=DATA_COL_START + 10, min_row=3, max_row=5)
    c3.add_data(c3_data, titles_from_data=True)
    c3.set_categories(c3_cats)
    c3.dataLabels               = DataLabelList()
    c3.dataLabels.showPercent   = True
    c3.dataLabels.showCatName   = True
    c3.dataLabels.showVal       = False
    c3.width = 14; c3.height = 14
    ws.add_chart(c3, 'B21')

    # --- Chart 4: Site Recurrence Bar ---
    c4 = BarChart()
    c4.type    = 'col'
    c4.grouping = 'clustered'
    c4.title   = 'Site Recurrence by Cluster'
    c4.y_axis.title = 'Site Count'
    c4.x_axis.title = 'Cluster'
    c4.style   = 10
    sr_end     = 3 + len(CLUSTER_ORDER) - 1
    c4_data    = Reference(ws, min_col=DATA_COL_START + 15, max_col=DATA_COL_START + 16,
                           min_row=2, max_row=sr_end)
    c4_cats    = Reference(ws, min_col=DATA_COL_START + 14, min_row=3, max_row=sr_end)
    c4.add_data(c4_data, titles_from_data=True)
    c4.set_categories(c4_cats)
    c4.width = 20; c4.height = 14
    ws.add_chart(c4, 'L21')

    ws.column_dimensions['A'].width = 2


# ──────────────────────────────────────────────────────────────────────────────
# Raw Data sheet & Named Ranges
# ──────────────────────────────────────────────────────────────────────────────
def _build_raw_sheet(wb: Workbook, df_str: pd.DataFrame) -> dict:
    """Write raw/normalized data and define dynamic Named Ranges for formulas."""
    ws = wb.create_sheet('Raw Data')
    ws.sheet_view.showGridLines = False

    # 1. Write headers and data
    cols = list(df_str.columns)
    for ci, col in enumerate(cols, start=1):
        _sc(ws, 1, ci, col, HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER)

    for ri, row in enumerate(df_str.itertuples(index=False), start=2):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row, start=1):
            col_name = cols[ci-1]
            if pd.isna(val): val = None
            
            # Preserve numeric types for specific helper columns
            if col_name in ['_UniqTT', '_SiteRepeat'] and val is not None:
                try: val = int(float(val))
                except: pass
            
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill; c.font = SMALL_FONT; c.border = THIN_BORDER; c.alignment = LEFT

            # MTTR formatting
            if col_name == '_E2EMTTR' and val is not None:
                if isinstance(val, datetime.timedelta):
                    c.value = _td_to_excel(val)
                c.number_format = TIME_FMT; c.alignment = CENTER

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 2. Define Dynamic Named Ranges
    last_row = len(df_str) + 1
    sname    = quote_sheetname('Raw Data')
    
    # Mapping of our internal logic names to the DataFrame/Excel column names
    name_map = {
        'Raw_Cluster':    '_Cluster',
        'Raw_Status':     '_WorkStatus',
        'Raw_Uniq':       '_UniqTT',
        'Raw_FO':         '_FO',
        'Raw_MTTR':       '_E2EMTTR',
        'Raw_Type':       '_TTType',
        'Raw_SubType':    '_SubType',
        'Raw_Action':     '_ActionType',
        'Raw_Repeat':     '_SiteRepeat'
    }

    refs = {}
    for dn, df_col in name_map.items():
        if df_col in cols:
            col_idx = cols.index(df_col) + 1
            let     = get_column_letter(col_idx)
            # Address format: 'Raw Data'!$B$2:$B$100
            addr    = f"{sname}!${let}$2:${let}${last_row}"
            refs[dn] = addr
            wb.defined_names.add(DefinedName(dn, attr_text=addr))
    
    return refs
